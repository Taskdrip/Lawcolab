from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify, Response
from flask_login import current_user
from app import db
from models import (CalendarEvent, CalendarEventAttendee, CourtDateHistory,
                    Project, ProjectAssignment, User,
                    EVENT_TYPE_COURT, EVENT_TYPE_MEETING, EVENT_TYPE_APPOINTMENT,
                    EVENT_TYPE_DEADLINE, EVENT_TYPE_OTHER,
                    EVENT_STATUS_UPCOMING, EVENT_STATUS_COMPLETED, EVENT_STATUS_CANCELLED)
from utils.decorators import simple_login_required
from datetime import datetime, date, timedelta
import calendar as cal_module
import io

calendar_bp = Blueprint('calendar', __name__)

EVENT_TYPES = [
    (EVENT_TYPE_COURT, 'Court Date'),
    (EVENT_TYPE_MEETING, 'Meeting'),
    (EVENT_TYPE_APPOINTMENT, 'Appointment'),
    (EVENT_TYPE_DEADLINE, 'Deadline'),
    (EVENT_TYPE_OTHER, 'Other'),
]

EVENT_STATUSES = [
    (EVENT_STATUS_UPCOMING, 'Upcoming'),
    (EVENT_STATUS_COMPLETED, 'Completed'),
    (EVENT_STATUS_CANCELLED, 'Cancelled'),
]

REMINDER_OPTIONS = [
    (0, 'At event time'),
    (15, '15 minutes before'),
    (30, '30 minutes before'),
    (60, '1 hour before'),
    (120, '2 hours before'),
    (1440, '1 day before'),
    (2880, '2 days before'),
    (10080, '1 week before'),
]

COURT_TYPES = [
    'Magistrate Court',
    'State High Court',
    'Federal High Court',
    'Court of Appeal',
    'Supreme Court',
    'Customary Court',
    'Sharia Court',
    'Other',
]

NIGERIAN_STATES = [
    'Abia', 'Adamawa', 'Akwa Ibom', 'Anambra', 'Bauchi', 'Bayelsa', 'Benue',
    'Borno', 'Cross River', 'Delta', 'Ebonyi', 'Edo', 'Ekiti', 'Enugu',
    'FCT - Abuja', 'Gombe', 'Imo', 'Jigawa', 'Kaduna', 'Kano', 'Katsina',
    'Kebbi', 'Kogi', 'Kwara', 'Lagos', 'Nasarawa', 'Niger', 'Ogun', 'Ondo',
    'Osun', 'Oyo', 'Plateau', 'Rivers', 'Sokoto', 'Taraba', 'Yobe', 'Zamfara',
]


def get_firm_events_query():
    """Base query for events visible to the current user."""
    if current_user.is_super_admin():
        return CalendarEvent.query
    q = CalendarEvent.query.filter_by(law_firm_id=current_user.law_firm_id)
    if current_user.is_client():
        q = q.join(CalendarEventAttendee).filter(CalendarEventAttendee.user_id == current_user.id)
    return q


def get_firm_projects():
    if current_user.is_admin() or current_user.is_super_admin():
        return Project.query.filter_by(law_firm_id=current_user.law_firm_id).all()
    return (Project.query
            .filter_by(law_firm_id=current_user.law_firm_id)
            .join(ProjectAssignment)
            .filter(ProjectAssignment.user_id == current_user.id)
            .all())


def get_firm_users():
    if not current_user.law_firm_id:
        return []
    return (User.query
            .filter_by(law_firm_id=current_user.law_firm_id, active=True)
            .all())


def _save_court_fields(event, form):
    """Copy court-specific fields from form into the event object."""
    event.court_jurisdiction = form.get('court_jurisdiction', '').strip() or None
    event.court_type = form.get('court_type', '').strip() or None
    event.court_address = form.get('court_address', '').strip() or None
    event.judge_name = form.get('judge_name', '').strip() or None


@calendar_bp.route('/')
@simple_login_required
def index():
    today = date.today()
    year = request.args.get('year', today.year, type=int)
    month = request.args.get('month', today.month, type=int)
    view = request.args.get('view', 'month')

    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    first_day = datetime(year, month, 1)
    last_day_num = cal_module.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59)

    events = (get_firm_events_query()
              .filter(CalendarEvent.start_datetime >= first_day,
                      CalendarEvent.start_datetime <= last_day)
              .order_by(CalendarEvent.start_datetime)
              .all())

    upcoming = (get_firm_events_query()
                .filter(CalendarEvent.start_datetime >= datetime.now(),
                        CalendarEvent.status == EVENT_STATUS_UPCOMING)
                .order_by(CalendarEvent.start_datetime)
                .limit(10)
                .all())

    cal = cal_module.monthcalendar(year, month)

    events_by_day = {}
    for ev in events:
        d = ev.start_datetime.day
        events_by_day.setdefault(d, []).append(ev)

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    month_name = first_day.strftime('%B %Y')

    week_offset = request.args.get('week', 0, type=int)
    today_dt = date.today()
    week_start = today_dt - timedelta(days=today_dt.weekday()) + timedelta(weeks=week_offset)
    week_days  = [week_start + timedelta(days=i) for i in range(7)]
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_end_dt   = datetime.combine(week_days[-1], datetime.max.time().replace(microsecond=0))

    week_events_raw = (get_firm_events_query()
                       .filter(CalendarEvent.start_datetime >= week_start_dt,
                               CalendarEvent.start_datetime <= week_end_dt)
                       .order_by(CalendarEvent.start_datetime)
                       .all())

    week_events_by_day = {}
    for ev in week_events_raw:
        d = ev.start_datetime.date()
        week_events_by_day.setdefault(d, []).append(ev)

    prev_week_offset = week_offset - 1
    next_week_offset = week_offset + 1
    week_label = (f"{week_days[0].strftime('%b %d')} – {week_days[-1].strftime('%b %d, %Y')}")

    events_by_date = {}
    for ev in events:
        d = ev.start_datetime.date()
        events_by_date.setdefault(d, []).append(ev)
    list_dates = sorted(events_by_date.keys())

    now_dt = datetime.now()
    due_today_ids = set()
    overdue_ids   = set()
    for ev in events:
        if ev.status == EVENT_STATUS_UPCOMING:
            if ev.start_datetime.date() == today_dt:
                due_today_ids.add(ev.id)
            elif ev.start_datetime < now_dt:
                overdue_ids.add(ev.id)

    return render_template('calendar/index.html',
                           cal=cal,
                           events=events,
                           events_by_day=events_by_day,
                           events_by_date=events_by_date,
                           list_dates=list_dates,
                           upcoming=upcoming,
                           year=year, month=month,
                           month_name=month_name,
                           today=today,
                           prev_month=prev_month, prev_year=prev_year,
                           next_month=next_month, next_year=next_year,
                           view=view,
                           week_days=week_days,
                           week_events_by_day=week_events_by_day,
                           week_offset=week_offset,
                           prev_week_offset=prev_week_offset,
                           next_week_offset=next_week_offset,
                           week_label=week_label,
                           due_today_ids=due_today_ids,
                           overdue_ids=overdue_ids)


@calendar_bp.route('/create', methods=['GET', 'POST'])
@simple_login_required
def create_event():
    if current_user.is_client():
        abort(403)

    projects = get_firm_projects()
    firm_users = get_firm_users()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Title is required.', 'danger')
            return render_template('calendar/create.html',
                                   event_types=EVENT_TYPES,
                                   reminder_options=REMINDER_OPTIONS,
                                   court_types=COURT_TYPES,
                                   nigerian_states=NIGERIAN_STATES,
                                   projects=projects,
                                   firm_users=firm_users)

        start_str = request.form.get('start_datetime', '')
        end_str = request.form.get('end_datetime', '')

        try:
            start_dt = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        except (ValueError, TypeError):
            flash('Invalid start date/time.', 'danger')
            return render_template('calendar/create.html',
                                   event_types=EVENT_TYPES,
                                   reminder_options=REMINDER_OPTIONS,
                                   court_types=COURT_TYPES,
                                   nigerian_states=NIGERIAN_STATES,
                                   projects=projects,
                                   firm_users=firm_users)

        end_dt = None
        if end_str:
            try:
                end_dt = datetime.strptime(end_str, '%Y-%m-%dT%H:%M')
            except (ValueError, TypeError):
                end_dt = None

        event = CalendarEvent()
        event.title = title
        event.description = request.form.get('description', '').strip()
        event.event_type = request.form.get('event_type', EVENT_TYPE_MEETING)
        event.status = EVENT_STATUS_UPCOMING
        event.start_datetime = start_dt
        event.end_datetime = end_dt
        event.all_day = 'all_day' in request.form
        event.location = request.form.get('location', '').strip()
        event.virtual_link = request.form.get('virtual_link', '').strip()
        event.notes = request.form.get('notes', '').strip()
        event.law_firm_id = current_user.law_firm_id
        event.created_by_id = current_user.id

        if event.event_type == EVENT_TYPE_COURT:
            _save_court_fields(event, request.form)

        proj_id = request.form.get('project_id', '')
        if proj_id:
            try:
                event.project_id = int(proj_id)
            except ValueError:
                event.project_id = None

        reminder_str = request.form.get('reminder_minutes', '60')
        try:
            event.reminder_minutes = int(reminder_str)
        except ValueError:
            event.reminder_minutes = 60

        db.session.add(event)
        db.session.flush()

        attendee_ids = request.form.getlist('attendee_ids')
        for uid in attendee_ids:
            if uid != current_user.id:
                att = CalendarEventAttendee(event_id=event.id, user_id=uid)
                db.session.add(att)
        creator_att = CalendarEventAttendee(event_id=event.id, user_id=current_user.id, rsvp_status='accepted')
        db.session.add(creator_att)

        db.session.commit()
        flash('Event created successfully!', 'success')
        return redirect(url_for('calendar.event_detail', event_id=event.id))

    default_date = request.args.get('date', '')
    default_dt = ''
    if default_date:
        try:
            d = datetime.strptime(default_date, '%Y-%m-%d')
            default_dt = d.strftime('%Y-%m-%dT09:00')
        except ValueError:
            pass

    return render_template('calendar/create.html',
                           event_types=EVENT_TYPES,
                           reminder_options=REMINDER_OPTIONS,
                           court_types=COURT_TYPES,
                           nigerian_states=NIGERIAN_STATES,
                           projects=projects,
                           firm_users=firm_users,
                           default_dt=default_dt)


@calendar_bp.route('/<int:event_id>')
@simple_login_required
def event_detail(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)
    can_edit = (event.created_by_id == current_user.id or
                current_user.is_admin() or current_user.is_super_admin())
    return render_template('calendar/detail.html', event=event, can_edit=can_edit)


@calendar_bp.route('/<int:event_id>/edit', methods=['GET', 'POST'])
@simple_login_required
def edit_event(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)
    if not (event.created_by_id == current_user.id or
            current_user.is_admin() or current_user.is_super_admin()):
        abort(403)

    projects = get_firm_projects()
    firm_users = get_firm_users()
    current_attendee_ids = [a.user_id for a in event.attendees]

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Title is required.', 'danger')
        else:
            start_str = request.form.get('start_datetime', '')
            end_str = request.form.get('end_datetime', '')

            try:
                event.start_datetime = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
            except (ValueError, TypeError):
                flash('Invalid start date/time.', 'danger')
                return render_template('calendar/edit.html', event=event,
                                       event_types=EVENT_TYPES, event_statuses=EVENT_STATUSES,
                                       reminder_options=REMINDER_OPTIONS,
                                       court_types=COURT_TYPES,
                                       nigerian_states=NIGERIAN_STATES,
                                       projects=projects, firm_users=firm_users,
                                       current_attendee_ids=current_attendee_ids)

            event.title = title
            event.description = request.form.get('description', '').strip()
            event.event_type = request.form.get('event_type', event.event_type)
            event.status = request.form.get('status', event.status)
            event.end_datetime = None
            if end_str:
                try:
                    event.end_datetime = datetime.strptime(end_str, '%Y-%m-%dT%H:%M')
                except (ValueError, TypeError):
                    pass
            event.all_day = 'all_day' in request.form
            event.location = request.form.get('location', '').strip()
            event.virtual_link = request.form.get('virtual_link', '').strip()
            event.notes = request.form.get('notes', '').strip()

            if event.event_type == EVENT_TYPE_COURT:
                _save_court_fields(event, request.form)
            else:
                event.court_jurisdiction = None
                event.court_type = None
                event.court_address = None
                event.judge_name = None

            proj_id = request.form.get('project_id', '')
            event.project_id = int(proj_id) if proj_id else None

            try:
                event.reminder_minutes = int(request.form.get('reminder_minutes', 60))
            except ValueError:
                event.reminder_minutes = 60

            CalendarEventAttendee.query.filter_by(event_id=event.id).delete()
            attendee_ids = request.form.getlist('attendee_ids')
            added = set()
            for uid in attendee_ids:
                if uid not in added:
                    att = CalendarEventAttendee(event_id=event.id, user_id=uid)
                    db.session.add(att)
                    added.add(uid)
            if current_user.id not in added:
                db.session.add(CalendarEventAttendee(event_id=event.id,
                                                     user_id=current_user.id,
                                                     rsvp_status='accepted'))

            db.session.commit()
            flash('Event updated.', 'success')
            return redirect(url_for('calendar.event_detail', event_id=event.id))

    return render_template('calendar/edit.html', event=event,
                           event_types=EVENT_TYPES, event_statuses=EVENT_STATUSES,
                           reminder_options=REMINDER_OPTIONS,
                           court_types=COURT_TYPES,
                           nigerian_states=NIGERIAN_STATES,
                           projects=projects, firm_users=firm_users,
                           current_attendee_ids=current_attendee_ids)


@calendar_bp.route('/<int:event_id>/delete', methods=['POST'])
@simple_login_required
def delete_event(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)
    if not (event.created_by_id == current_user.id or
            current_user.is_admin() or current_user.is_super_admin()):
        abort(403)
    db.session.delete(event)
    db.session.commit()
    flash('Event deleted.', 'success')
    return redirect(url_for('calendar.index'))


@calendar_bp.route('/<int:event_id>/status', methods=['POST'])
@simple_login_required
def update_status(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)
    new_status = request.form.get('status')
    if new_status in (EVENT_STATUS_UPCOMING, EVENT_STATUS_COMPLETED, EVENT_STATUS_CANCELLED):
        event.status = new_status
        db.session.commit()
        flash('Status updated.', 'success')
    return redirect(url_for('calendar.event_detail', event_id=event_id))


# ── Court date history ────────────────────────────────────────────────────────

def _can_edit_event(event):
    """Return True if the current user may create/edit/delete records on this event."""
    return (event.created_by_id == current_user.id or
            current_user.is_admin() or current_user.is_super_admin())


@calendar_bp.route('/<int:event_id>/history/add', methods=['POST'])
@simple_login_required
def add_court_history(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)
    # Only the event creator, firm admins, and super-admins may log history
    if not _can_edit_event(event):
        abort(403)
    # History only makes sense on court-date events
    if event.event_type != EVENT_TYPE_COURT:
        abort(400)

    hearing_date_str = request.form.get('hearing_date', '').strip()
    try:
        hearing_date = datetime.strptime(hearing_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        flash('Invalid date for history entry.', 'danger')
        return redirect(url_for('calendar.event_detail', event_id=event_id))

    entry = CourtDateHistory(
        event_id=event_id,
        hearing_date=hearing_date,
        outcome=request.form.get('outcome', '').strip() or None,
        court_notes=request.form.get('court_notes', '').strip() or None,
        recorded_by_id=current_user.id,
    )
    db.session.add(entry)
    db.session.commit()
    flash('Court history entry added.', 'success')
    return redirect(url_for('calendar.event_detail', event_id=event_id))


@calendar_bp.route('/history/<int:history_id>/delete', methods=['POST'])
@simple_login_required
def delete_court_history(history_id):
    entry = CourtDateHistory.query.get_or_404(history_id)
    event = CalendarEvent.query.get_or_404(entry.event_id)
    _check_access(event)
    # Admins/super-admins can delete any entry; others can only delete their own
    if not (current_user.is_admin() or current_user.is_super_admin() or
            entry.recorded_by_id == current_user.id):
        abort(403)
    db.session.delete(entry)
    db.session.commit()
    flash('History entry deleted.', 'success')
    return redirect(url_for('calendar.event_detail', event_id=entry.event_id))


# ── Export: Excel ─────────────────────────────────────────────────────────────

@calendar_bp.route('/export/excel')
@simple_login_required
def export_excel():
    """Export all court-date events (with history) for the firm as an Excel workbook."""
    if current_user.is_client():
        abort(403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    events = (get_firm_events_query()
              .filter(CalendarEvent.event_type == EVENT_TYPE_COURT)
              .order_by(CalendarEvent.start_datetime)
              .all())

    wb = openpyxl.Workbook()

    # ── Sheet 1: Upcoming court dates ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Court Dates'

    header_fill = PatternFill('solid', fgColor='0A1847')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(border_style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Case / Event Title', 'Linked Project', 'Date', 'Time',
        'Jurisdiction (State)', 'Court Type', 'Court Address',
        'Judge / Magistrate', 'Status', 'Description', 'Internal Notes',
    ]
    col_widths = [35, 25, 14, 10, 20, 22, 35, 25, 12, 40, 40]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 28

    for row_idx, ev in enumerate(events, 2):
        row_fill = PatternFill('solid', fgColor='F9FAFB') if row_idx % 2 == 0 else None
        values = [
            ev.title,
            ev.project.title if ev.project else '',
            ev.start_datetime.strftime('%Y-%m-%d'),
            '' if ev.all_day else ev.start_datetime.strftime('%I:%M %p'),
            ev.court_jurisdiction or '',
            ev.court_type or '',
            ev.court_address or '',
            ev.judge_name or '',
            ev.status.title(),
            ev.description or '',
            ev.notes or '',
        ]
        for ci, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
        ws1.row_dimensions[row_idx].height = 22

    # ── Sheet 2: Previous hearing history ─────────────────────────────────
    ws2 = wb.create_sheet('Hearing History')
    h2 = ['Case / Event Title', 'Linked Project', 'Hearing Date', 'Outcome / Result', 'Court Notes', 'Recorded By']
    w2 = [35, 25, 14, 25, 55, 22]
    for ci, (h, w) in enumerate(zip(h2, w2), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 28

    row_idx2 = 2
    for ev in events:
        for h in ev.court_history:
            row_fill = PatternFill('solid', fgColor='F9FAFB') if row_idx2 % 2 == 0 else None
            values = [
                ev.title,
                ev.project.title if ev.project else '',
                h.hearing_date.strftime('%Y-%m-%d'),
                h.outcome or '',
                h.court_notes or '',
                h.recorded_by.full_name if h.recorded_by else '',
            ]
            for ci, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx2, column=ci, value=val)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
            ws2.row_dimensions[row_idx2].height = 22
            row_idx2 += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"court_calendar_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ── Export: iCal ──────────────────────────────────────────────────────────────

@calendar_bp.route('/export/ical')
@simple_login_required
def export_ical():
    """Export upcoming events as an .ics file compatible with Google Calendar, Outlook, Apple Calendar."""
    from icalendar import Calendar, Event as IEvent
    import pytz

    events = (get_firm_events_query()
              .filter(CalendarEvent.start_datetime >= datetime.now(),
                      CalendarEvent.status == EVENT_STATUS_UPCOMING)
              .order_by(CalendarEvent.start_datetime)
              .all())

    cal = Calendar()
    cal.add('prodid', '-//LawFirmOS//Court Calendar//EN')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    cal.add('x-wr-calname', 'LawFirmOS – Court Dates')
    cal.add('x-wr-timezone', 'Africa/Lagos')

    tz = pytz.timezone('Africa/Lagos')

    for ev in events:
        iev = IEvent()
        iev.add('uid', f'lawfirmos-event-{ev.id}@lawcolab')
        iev.add('summary', ev.title)

        start = tz.localize(ev.start_datetime) if ev.start_datetime.tzinfo is None else ev.start_datetime
        iev.add('dtstart', start)

        if ev.end_datetime:
            end = tz.localize(ev.end_datetime) if ev.end_datetime.tzinfo is None else ev.end_datetime
            iev.add('dtend', end)

        description_parts = []
        if ev.description:
            description_parts.append(ev.description)
        if ev.event_type == EVENT_TYPE_COURT:
            if ev.court_jurisdiction:
                description_parts.append(f'Jurisdiction: {ev.court_jurisdiction}')
            if ev.court_type:
                description_parts.append(f'Court: {ev.court_type}')
            if ev.judge_name:
                description_parts.append(f'Judge/Magistrate: {ev.judge_name}')
        if description_parts:
            iev.add('description', '\n'.join(description_parts))

        location_parts = []
        if ev.court_address:
            location_parts.append(ev.court_address)
        elif ev.location:
            location_parts.append(ev.location)
        if location_parts:
            iev.add('location', ' | '.join(location_parts))

        iev.add('status', 'CONFIRMED')
        iev.add('dtstamp', datetime.now(pytz.utc))
        cal.add_component(iev)

    ical_bytes = cal.to_ical()
    filename = f"lawfirmos_calendar_{date.today().strftime('%Y%m%d')}.ics"
    return Response(
        ical_bytes,
        mimetype='text/calendar',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ── Export: iCal for single event ─────────────────────────────────────────────

@calendar_bp.route('/<int:event_id>/export/ical')
@simple_login_required
def export_event_ical(event_id):
    """Download a single event as .ics."""
    from icalendar import Calendar, Event as IEvent
    import pytz

    event = CalendarEvent.query.get_or_404(event_id)
    _check_access(event)

    tz = pytz.timezone('Africa/Lagos')
    cal = Calendar()
    cal.add('prodid', '-//LawFirmOS//Court Calendar//EN')
    cal.add('version', '2.0')

    iev = IEvent()
    iev.add('uid', f'lawfirmos-event-{event.id}@lawcolab')
    iev.add('summary', event.title)
    start = tz.localize(event.start_datetime) if event.start_datetime.tzinfo is None else event.start_datetime
    iev.add('dtstart', start)
    if event.end_datetime:
        end = tz.localize(event.end_datetime) if event.end_datetime.tzinfo is None else event.end_datetime
        iev.add('dtend', end)

    desc_parts = []
    if event.description:
        desc_parts.append(event.description)
    if event.event_type == EVENT_TYPE_COURT:
        for label, val in [('Jurisdiction', event.court_jurisdiction),
                           ('Court', event.court_type),
                           ('Judge/Magistrate', event.judge_name)]:
            if val:
                desc_parts.append(f'{label}: {val}')
    if desc_parts:
        iev.add('description', '\n'.join(desc_parts))

    location = event.court_address or event.location
    if location:
        iev.add('location', location)

    iev.add('status', 'CONFIRMED')
    iev.add('dtstamp', datetime.now(pytz.utc))
    cal.add_component(iev)

    filename = f"event_{event.id}.ics"
    return Response(
        cal.to_ical(),
        mimetype='text/calendar',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@calendar_bp.route('/history/<int:history_id>/edit', methods=['GET', 'POST'])
@simple_login_required
def edit_court_history(history_id):
    entry = CourtDateHistory.query.get_or_404(history_id)
    event = CalendarEvent.query.get_or_404(entry.event_id)
    _check_access(event)
    if not (current_user.is_admin() or current_user.is_super_admin() or
            entry.recorded_by_id == current_user.id):
        abort(403)

    if request.method == 'POST':
        hearing_date_str = request.form.get('hearing_date', '').strip()
        try:
            entry.hearing_date = datetime.strptime(hearing_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid date.', 'danger')
            return redirect(url_for('calendar.event_detail', event_id=event.id))

        entry.outcome = request.form.get('outcome', '').strip() or None
        entry.court_notes = request.form.get('court_notes', '').strip() or None
        db.session.commit()
        flash('History entry updated.', 'success')
        return redirect(url_for('calendar.event_detail', event_id=event.id))

    return render_template('calendar/edit_history.html', entry=entry, event=event)


@calendar_bp.route('/court-docket')
@simple_login_required
def court_docket():
    """Dedicated court docket — all court-date events with filters and history summary."""
    if current_user.is_client():
        abort(403)

    # Filter params
    f_jurisdiction = request.args.get('jurisdiction', '').strip()
    f_court_type   = request.args.get('court_type', '').strip()
    f_judge        = request.args.get('judge', '').strip()
    f_status       = request.args.get('status', '').strip()
    f_project      = request.args.get('project_id', '', type=str).strip()

    q = (get_firm_events_query()
         .filter(CalendarEvent.event_type == EVENT_TYPE_COURT))

    if f_jurisdiction:
        q = q.filter(CalendarEvent.court_jurisdiction == f_jurisdiction)
    if f_court_type:
        q = q.filter(CalendarEvent.court_type == f_court_type)
    if f_judge:
        q = q.filter(CalendarEvent.judge_name.ilike(f'%{f_judge}%'))
    if f_status:
        q = q.filter(CalendarEvent.status == f_status)
    if f_project:
        try:
            q = q.filter(CalendarEvent.project_id == int(f_project))
        except ValueError:
            pass

    events = q.order_by(CalendarEvent.start_datetime).all()

    # Stats
    total = len(events)
    upcoming_count  = sum(1 for e in events if e.status == EVENT_STATUS_UPCOMING)
    completed_count = sum(1 for e in events if e.status == EVENT_STATUS_COMPLETED)
    cancelled_count = sum(1 for e in events if e.status == EVENT_STATUS_CANCELLED)

    # Unique values for filter dropdowns (from all firm court events, not filtered subset)
    all_court_q = (get_firm_events_query()
                   .filter(CalendarEvent.event_type == EVENT_TYPE_COURT))
    all_court_events = all_court_q.all()

    jurisdictions = sorted({e.court_jurisdiction for e in all_court_events if e.court_jurisdiction})
    judges        = sorted({e.judge_name for e in all_court_events if e.judge_name})
    projects      = get_firm_projects()

    return render_template('calendar/court_docket.html',
                           events=events,
                           total=total,
                           upcoming_count=upcoming_count,
                           completed_count=completed_count,
                           cancelled_count=cancelled_count,
                           jurisdictions=jurisdictions,
                           judges=judges,
                           court_types=COURT_TYPES,
                           event_statuses=EVENT_STATUSES,
                           projects=projects,
                           f_jurisdiction=f_jurisdiction,
                           f_court_type=f_court_type,
                           f_judge=f_judge,
                           f_status=f_status,
                           f_project=f_project,
                           today=date.today())


@calendar_bp.route('/upcoming')
@simple_login_required
def upcoming_events():
    events = (get_firm_events_query()
              .filter(CalendarEvent.start_datetime >= datetime.now(),
                      CalendarEvent.status == EVENT_STATUS_UPCOMING)
              .order_by(CalendarEvent.start_datetime)
              .all())
    return render_template('calendar/upcoming.html', events=events, today=date.today())


def _check_access(event):
    if current_user.is_super_admin():
        return
    if event.law_firm_id != current_user.law_firm_id:
        abort(403)
    if current_user.is_client():
        ids = [a.user_id for a in event.attendees]
        if current_user.id not in ids:
            abort(403)


# ── JSON API for client-side reminder system ──────────────────────────────────

@calendar_bp.route('/api/reminders')
@simple_login_required
def api_reminders():
    now = datetime.now()
    window_end = now + timedelta(days=7)

    events = (get_firm_events_query()
              .filter(CalendarEvent.start_datetime >= now,
                      CalendarEvent.start_datetime <= window_end,
                      CalendarEvent.status == EVENT_STATUS_UPCOMING)
              .order_by(CalendarEvent.start_datetime)
              .all())

    result = []
    for ev in events:
        reminder_mins = ev.reminder_minutes if ev.reminder_minutes is not None else 60
        reminder_fire_ts = int(
            (ev.start_datetime - timedelta(minutes=reminder_mins)).timestamp() * 1000
        )
        event_ts = int(ev.start_datetime.timestamp() * 1000)

        result.append({
            'id': ev.id,
            'title': ev.title,
            'event_type': ev.event_type,
            'type_label': ev.type_label,
            'type_color': ev.type_color,
            'type_icon': ev.type_icon,
            'start_ts': event_ts,
            'start_display': ev.start_datetime.strftime('%A, %b %d, %Y at %I:%M %p'),
            'start_date': ev.start_datetime.strftime('%Y-%m-%d'),
            'all_day': ev.all_day,
            'location': ev.location or '',
            'virtual_link': ev.virtual_link or '',
            'description': ev.description or '',
            'notes': ev.notes or '',
            'reminder_minutes': reminder_mins,
            'reminder_fire_ts': reminder_fire_ts,
            'detail_url': url_for('calendar.event_detail', event_id=ev.id),
            'attendees': [a.user.full_name for a in ev.attendees],
        })

    return jsonify({'events': result, 'server_ts': int(now.timestamp() * 1000)})


@calendar_bp.route('/api/events')
@simple_login_required
def api_events():
    today = date.today()
    year = request.args.get('year', today.year, type=int)
    month = request.args.get('month', today.month, type=int)

    first_day = datetime(year, month, 1)
    last_day_num = cal_module.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59)

    events = (get_firm_events_query()
              .filter(CalendarEvent.start_datetime >= first_day,
                      CalendarEvent.start_datetime <= last_day)
              .order_by(CalendarEvent.start_datetime)
              .all())

    result = []
    for ev in events:
        result.append({
            'id': ev.id,
            'title': ev.title,
            'event_type': ev.event_type,
            'type_label': ev.type_label,
            'type_color': ev.type_color,
            'type_icon': ev.type_icon,
            'start_ts': int(ev.start_datetime.timestamp() * 1000),
            'start_day': ev.start_datetime.day,
            'start_display': ev.start_datetime.strftime('%b %d, %Y %I:%M %p'),
            'all_day': ev.all_day,
            'status': ev.status,
            'location': ev.location or '',
            'reminder_minutes': ev.reminder_minutes,
            'detail_url': url_for('calendar.event_detail', event_id=ev.id),
        })

    return jsonify({'events': result})
